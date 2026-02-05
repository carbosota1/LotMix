import os
import re
import csv
import base64
from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
TZ_RD = ZoneInfo("America/Santo_Domingo")

BASE_URL = "https://www.loteriadominicana.com.do/Lottery/National"

OUT_CSV  = "La nacional history.csv"
OUT_XLSX = "La nacional history.xlsx"

CHECKPOINT_DIR = "checkpoints"
CHECKPOINT_EVERY_DAYS = 15

TARGET_DRAWS = {
    "Loteria Nacional- Gana Más",
    "Loteria Nacional- Noche",
}

DEFAULT_START_DATE = date(2021, 1, 1)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome Safari"
    )
}

CSV_COLUMNS = [
    "fecha", "sorteo", "primero", "segundo", "tercero",
    "source", "url_key", "scraped_at_rd"
]

# =========================
# UTILIDADES
# =========================
def ensure_dirs():
    os.makedirs(CHECKPOINT_DIR, exist_ok=True)

def today_rd() -> date:
    return datetime.now(TZ_RD).date()

def z2(x: str) -> str:
    """
    Normaliza a 2 dígitos SIN perder 00.
    """
    s = str(x).strip()

    # ya está en 2 dígitos => respeta (00, 07, 84, etc.)
    if re.fullmatch(r"\d{2}", s):
        return s

    m = re.search(r"\d+", s)
    if not m:
        return ""
    return m.group(0).zfill(2)

def encode_d_param(d: date) -> str:
    """
    Genera el ?d= que usa loteriadominicana.com.do

    Patrón:
    - ddmmyyyy (ej: 07062024)
    - invertir -> 42026070
    - decimal -> HEX (uppercase)
    - base64(HEX)
    """
    ddmmyyyy = d.strftime("%d%m%Y")
    rev = ddmmyyyy[::-1]
    hx = format(int(rev), "X")
    return base64.b64encode(hx.encode("utf-8")).decode("utf-8")

def build_url_for_date(d: date) -> tuple[str, str]:
    key = encode_d_param(d)
    return f"{BASE_URL}?d={key}", key

def parse_date_from_text(txt: str) -> date | None:
    m = re.search(r"\b(\d{2})-(\d{2})-(\d{4})\b", txt)
    if not m:
        return None
    dd, mm, yyyy = map(int, m.groups())
    return date(yyyy, mm, dd)

def load_existing(csv_path: str):
    keys = set()
    last_date = None

    if not os.path.exists(csv_path):
        return keys, last_date

    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            k = (r.get("fecha", ""), r.get("sorteo", ""))
            keys.add(k)
            try:
                d = datetime.strptime(r.get("fecha", ""), "%Y-%m-%d").date()
                if last_date is None or d > last_date:
                    last_date = d
            except Exception:
                pass

    return keys, last_date

def append_rows(csv_path: str, new_rows: list[dict]):
    file_exists = os.path.exists(csv_path)
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        if not file_exists:
            writer.writeheader()
        for r in new_rows:
            writer.writerow(r)

def write_checkpoint_snapshot(src_csv: str, start_d: date, end_d: date):
    if not os.path.exists(src_csv):
        return
    snap_name = f"La nacional history__{start_d.isoformat()}__{end_d.isoformat()}.csv"
    snap_path = os.path.join(CHECKPOINT_DIR, snap_name)
    with open(src_csv, "r", encoding="utf-8") as fr, open(snap_path, "w", encoding="utf-8") as fw:
        fw.write(fr.read())
    print(f"[checkpoint] Snapshot creado: {snap_path}")

def daterange(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)

def get_result(draw: str, date: str) -> tuple[str, str, str]:
    # TODO: tu scraping real aquí
    raise NotImplementedError


# =========================
# EXPORT XLSX (Excel friendly)
# =========================
def export_xlsx_from_csv(csv_path: str, xlsx_path: str):
    if not os.path.exists(csv_path):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "La nacional history"

    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        wb.save(xlsx_path)
        return

    for row in rows:
        ws.append(row)

    header = rows[0]
    col_map = {name: (idx + 1) for idx, name in enumerate(header)}

    # Asegura que Excel no quite ceros en 00, 07, etc.
    for col_name in ("primero", "segundo", "tercero"):
        if col_name in col_map:
            c = col_map[col_name]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=c)
                cell.number_format = "@"  # texto

    # Auto-ajuste de columnas
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    wb.save(xlsx_path)
    print(f"[xlsx] Exportado: {xlsx_path}")

# =========================
# SCRAPER
# =========================
def fetch_day_results(d: date) -> list[dict]:
    url, url_key = build_url_for_date(d)
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()

    soup = BeautifulSoup(r.text, "html.parser")

    out = []
    for h4 in soup.find_all("h4"):
        title = (h4.get_text(strip=True) or "")
        if title not in TARGET_DRAWS:
            continue

        # Buscar contenedor cercano con bolas
        container = h4.parent
        for _ in range(6):
            if container is None:
                break
            if container.find(class_=re.compile(r"result-item-ball-content|ball")):
                break
            container = container.parent
        if not container:
            continue

        # Extraer bolas desde HTML (más confiable)
        balls = container.select("div.ball span")
        nums = [z2(b.get_text(strip=True)) for b in balls if b.get_text(strip=True)]

        # Fallback si cambia el HTML
        if len(nums) < 3:
            txt_nums = re.findall(r"\b\d{1,2}\b", container.get_text(" ", strip=True))
            nums = [z2(x) for x in txt_nums][:3]

        if len(nums) < 3:
            continue

        block_text = container.get_text(" ", strip=True)
        block_date = parse_date_from_text(block_text) or d

        out.append({
            "fecha": block_date.isoformat(),
            "sorteo": title,
            "primero": nums[0],
            "segundo": nums[1],
            "tercero": nums[2],
            "source": "LD",
            "url_key": url_key,
            "scraped_at_rd": datetime.now(TZ_RD).strftime("%Y-%m-%d %H:%M:%S"),
        })

    return out

def main():
    ensure_dirs()

    existing_keys, last_date = load_existing(OUT_CSV)

    if last_date is None:
        start_date = DEFAULT_START_DATE
        print(f"[init] CSV no existe o vacío. Iniciando desde {start_date}")
    else:
        start_date = last_date + timedelta(days=1)
        print(f"[resume] Última fecha en CSV: {last_date} -> Continuando desde {start_date}")

    end_date = today_rd()
    if start_date > end_date:
        print("[ok] No hay nada nuevo que extraer.")
        export_xlsx_from_csv(OUT_CSV, OUT_XLSX)
        return

    print(f"[run] Extrayendo desde {start_date} hasta {end_date}")

    checkpoint_window_start = start_date
    days_in_window = 0
    any_new = False

    for d in daterange(start_date, end_date):
        try:
            day_rows = fetch_day_results(d)

            new_rows = []
            for row in day_rows:
                k = (row["fecha"], row["sorteo"])
                if k in existing_keys:
                    continue
                existing_keys.add(k)
                new_rows.append(row)

            if new_rows:
                append_rows(OUT_CSV, new_rows)
                any_new = True
                print(f"[+]{d} -> {len(new_rows)} registros nuevos")
            else:
                print(f"[=]{d} -> sin nuevos (o aún no publicados)")

        except Exception as e:
            print(f"[error] {d} -> {e}")

        days_in_window += 1

        if days_in_window >= CHECKPOINT_EVERY_DAYS:
            write_checkpoint_snapshot(OUT_CSV, checkpoint_window_start, d)
            checkpoint_window_start = d + timedelta(days=1)
            days_in_window = 0

    if days_in_window > 0:
        write_checkpoint_snapshot(OUT_CSV, checkpoint_window_start, end_date)

    export_xlsx_from_csv(OUT_CSV, OUT_XLSX)

    if any_new:
        print("[done] Listo. CSV actualizado y XLSX generado.")
    else:
        print("[done] No hubo filas nuevas. XLSX regenerado para visualización.")

if __name__ == "__main__":
    main()
